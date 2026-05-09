import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

src = Path('financial_data.json')
out = Path('financial_report.xlsx')

with src.open('r', encoding='utf-8-sig') as f:
    data = json.load(f)

wb = Workbook()
ws_detail = wb.active
ws_detail.title = 'financial_report'
ws_month = wb.create_sheet('每月統計')
ws_total = wb.create_sheet('總統計')

headers = ['月份','日期','類型','分類','項目','金額']
ws_detail.append(headers)
for c in ws_detail[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='D9E1F2')

monthly_summary = []
for m in data['months']:
    month = m['month']
    income = 0
    expense = 0
    for d in m['details']:
        amt = float(d['amount'])
        ws_detail.append([month, d['date'], d['type'], d['category'], d['item'], amt])
        if d['type'] == 'income':
            income += amt
        else:
            expense += amt
    monthly_summary.append((month, income, expense, income-expense))

for col in ['A','B','C','D','E','F']:
    ws_detail.column_dimensions[col].width = 18
for r in ws_detail.iter_rows(min_row=2, max_col=6):
    r[5].number_format = '#,##0.00'

ws_month.append(['月份','收入','支出','淨利'])
for c in ws_month[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='FCE4D6')

for row in monthly_summary:
    ws_month.append(list(row))
for col in ['A','B','C','D']:
    ws_month.column_dimensions[col].width = 18
for r in ws_month.iter_rows(min_row=2, max_col=4):
    for c in r[1:]:
        c.number_format = '#,##0.00'

sum_income = sum(x[1] for x in monthly_summary)
sum_expense = sum(x[2] for x in monthly_summary)
sum_profit = sum(x[3] for x in monthly_summary)

ws_total.append(['項目','金額'])
for c in ws_total[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='E2F0D9')
ws_total.append(['總收入', sum_income])
ws_total.append(['總支出', sum_expense])
ws_total.append(['總淨利', sum_profit])
for col in ['A','B']:
    ws_total.column_dimensions[col].width = 20
for r in ws_total.iter_rows(min_row=2, max_col=2):
    r[1].number_format = '#,##0.00'

for ws in [ws_detail, ws_month, ws_total]:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')

wb.save(out)
print(f'Wrote: {out.resolve()}')
